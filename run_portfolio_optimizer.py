#!/usr/bin/env python3
"""
run_portfolio_optimizer.py  (zerobot)

Lädt alle EAR-Configs, führt Portfolio-Simulation durch und wählt das beste
Portfolio per Greedy-Algorithmus. Schreibt active_strategies in settings.json.

Aufruf:
  python3 run_portfolio_optimizer.py              # interaktiv
  python3 run_portfolio_optimizer.py --auto-write # automatisch (Scheduler)
  python3 run_portfolio_optimizer.py --replot     # Replot aktives Portfolio
"""
import contextlib
import io
import os
import sys
import json
import argparse
import pandas as pd
from datetime import date, timedelta
from tqdm import tqdm

PROJECT_ROOT  = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, os.path.join(PROJECT_ROOT, 'src'))

CONFIGS_DIR   = os.path.join(PROJECT_ROOT, 'src', 'zerobot', 'strategy', 'configs')
SETTINGS_PATH = os.path.join(PROJECT_ROOT, 'settings.json')

B  = '\033[1;37m'
G  = '\033[0;32m'
Y  = '\033[1;33m'
R  = '\033[0;31m'
NC = '\033[0m'

DEFAULT_LOOKBACK_DAYS = 730

BOT_NAME = 'zerobot'


def _scan_configs() -> list:
    if not os.path.isdir(CONFIGS_DIR):
        return []
    return sorted([
        os.path.join(CONFIGS_DIR, f)
        for f in os.listdir(CONFIGS_DIR)
        if f.endswith('.json')
    ])


def _build_strategies_data(config_files: list, start_date: str, end_date: str) -> dict:
    from zerobot.analysis.backtester import load_data
    strategies_data = {}
    for path in tqdm(config_files, desc='Lade Configs & Daten'):
        fname = os.path.basename(path)
        try:
            with open(path) as f:
                config = json.load(f)
            market    = config.get('market', {})
            symbol    = market.get('symbol', '')
            timeframe = market.get('timeframe', '')
            htf       = market.get('htf')
            if not symbol or not timeframe:
                continue
            data = load_data(symbol, timeframe, start_date, end_date)
            if data is None or data.empty or len(data) < 50:
                print(f"  {Y}Übersprungen (keine Daten): {fname}{NC}")
                continue
            strategies_data[fname] = {
                'symbol':      symbol,
                'timeframe':   timeframe,
                'data':        data,
                'smc_params':  config.get('strategy', {}),
                'risk_params': config.get('risk', {}),
                'htf':         htf,
            }
        except Exception as e:
            print(f"  {Y}Fehler bei {fname}: {e}{NC}")
    return strategies_data


def _write_to_settings(portfolio_files: list, strategies_data: dict) -> None:
    with open(SETTINGS_PATH) as f:
        settings = json.load(f)
    existing     = settings.get('live_trading_settings', {}).get('active_strategies', [])
    existing_map = {(s.get('symbol'), s.get('timeframe')): s for s in existing}
    new_strategies = []
    for fname in portfolio_files:
        sd        = strategies_data.get(fname, {})
        symbol    = sd.get('symbol', '')
        timeframe = sd.get('timeframe', '')
        if not symbol or not timeframe:
            continue
        base  = existing_map.get((symbol, timeframe), {})
        entry = {**base, 'symbol': symbol, 'timeframe': timeframe, 'active': True}
        new_strategies.append(entry)
    lt = settings.setdefault('live_trading_settings', {})
    lt['active_strategies']          = new_strategies
    lt['use_auto_optimizer_results'] = True
    with open(SETTINGS_PATH, 'w') as f:
        json.dump(settings, f, indent=4)


def _get_telegram_creds():
    try:
        with open(os.path.join(PROJECT_ROOT, 'secret.json')) as f:
            s = json.load(f)
        tg = s.get('telegram', {})
        t, c = tg.get('bot_token', ''), tg.get('chat_id', '')
        return (t, c) if t and c else (None, None)
    except Exception:
        return None, None


def _send_telegram(msg):
    token, chat = _get_telegram_creds()
    if not token:
        return
    try:
        import requests
        requests.post(f'https://api.telegram.org/bot{token}/sendMessage',
                      data={'chat_id': chat, 'text': msg}, timeout=10)
    except Exception:
        pass


def _send_telegram_doc(fpath, caption=''):
    token, chat = _get_telegram_creds()
    if not token:
        return
    try:
        import requests
        with open(fpath, 'rb') as fh:
            requests.post(f'https://api.telegram.org/bot{token}/sendDocument',
                          data={'chat_id': chat, 'caption': caption},
                          files={'document': fh}, timeout=30)
    except Exception:
        pass


PAIR_COLORS = [
    '#f59e0b', '#8b5cf6', '#ec4899', '#14b8a6',
    '#f97316', '#84cc16', '#06b6d4', '#a78bfa',
]


def generate_equity_html(final, capital, start_date, end_date, labels, portfolio_files=None, strategies_data=None):
    try:
        import plotly.graph_objects as go
        from plotly.subplots import make_subplots
    except ImportError:
        print(f'  {Y}plotly nicht installiert — Chart übersprungen.{NC}')
        return None

    eq_df   = final.get('equity_curve')
    trades  = final.get('trade_history', [])
    if eq_df is None or (hasattr(eq_df, 'empty') and eq_df.empty):
        return None

    # equity_curve has timestamp as index and column
    if 'timestamp' in eq_df.columns:
        eq_times = [str(t) for t in eq_df['timestamp']]
    else:
        eq_times = [str(t) for t in eq_df.index]
    eq_vals = [float(v) for v in eq_df['equity']]

    pnl  = final.get('total_pnl_pct', 0)
    dd   = final.get('max_drawdown_pct', 0)
    wr   = final.get('win_rate', 0)
    n    = final.get('trade_count', 0)
    eq   = final.get('end_capital', eq_vals[-1] if eq_vals else capital)
    sign = '+' if pnl >= 0 else ''
    pairs_str = ', '.join(labels) if labels else BOT_NAME

    title = (f"{BOT_NAME} Portfolio — {pairs_str} | "
             f"Trades: {n} | WR: {wr:.1f}% | "
             f"PnL: {sign}{pnl:.1f}% | Equity: {eq:.2f} USDT | MaxDD: {dd:.1f}%")

    fig = make_subplots(specs=[[{"secondary_y": True}]])

    # Per-strategy equity overlay (primary y-axis, thin colored lines)
    if trades and strategies_data:
        strategy_pairs = []
        for fname, sd in strategies_data.items():
            sym = sd.get('symbol', '')
            tf  = sd.get('timeframe', '')
            if sym and tf:
                strategy_pairs.append((sym, tf))
        for idx, (sym, tf) in enumerate(strategy_pairs):
            pair_trades = sorted(
                [t for t in trades if t.get('symbol') == sym and t.get('timeframe') == tf],
                key=lambda t: t.get('entry_time', ''),
            )
            if not pair_trades:
                continue
            peq    = capital
            ptimes = [pair_trades[0].get('entry_time', '')[:19]]
            pvals  = [peq]
            for t in pair_trades:
                peq += t.get('pnl', 0)
                ptimes.append(t.get('ts', t.get('entry_time', ''))[:19])
                pvals.append(round(peq, 2))
            label = f"{sym.split('/')[0]}/{tf}"
            fig.add_trace(go.Scatter(
                x=ptimes, y=pvals, mode='lines', name=label,
                line=dict(color=PAIR_COLORS[idx % len(PAIR_COLORS)], width=1),
                opacity=0.55,
            ), secondary_y=False)

    # Start capital reference line
    fig.add_hline(y=capital,
                  line=dict(color='rgba(100,100,100,0.35)', width=1, dash='dash'),
                  annotation_text=f'Start {capital:.0f} USDT',
                  annotation_position='top left')

    # WIN / LOSS markers on portfolio equity (secondary y-axis)
    win_x, win_y   = [], []
    loss_x, loss_y = [], []
    if trades and eq_df is not None:
        # build a lookup: timestamp string → equity value
        eq_idx = eq_df.copy()
        eq_idx.index = pd.to_datetime(eq_idx.index, utc=True, errors='coerce')
        for t in trades:
            ts_str = t.get('ts', '')
            try:
                ts_dt = pd.to_datetime(ts_str, utc=True, errors='coerce')
                if pd.isna(ts_dt):
                    continue
                # find nearest equity value
                idx_pos = eq_idx.index.get_indexer([ts_dt], method='nearest')[0]
                eq_val  = float(eq_idx['equity'].iloc[idx_pos])
            except Exception:
                continue
            if t.get('pnl', 0) > 0:
                win_x.append(ts_str[:19])
                win_y.append(eq_val)
            else:
                loss_x.append(ts_str[:19])
                loss_y.append(eq_val)

    # Portfolio equity curve (secondary y-axis, blue)
    fig.add_trace(go.Scatter(
        x=eq_times, y=eq_vals, mode='lines', name='Portfolio Equity',
        line=dict(color='#2563eb', width=2), opacity=0.75,
    ), secondary_y=True)

    if win_x:
        fig.add_trace(go.Scatter(
            x=win_x, y=win_y, mode='markers', name='Exit TP ✓',
            marker=dict(color='#22d3ee', symbol='circle', size=10,
                        line=dict(width=1, color='#0e7490')),
        ), secondary_y=True)
    if loss_x:
        fig.add_trace(go.Scatter(
            x=loss_x, y=loss_y, mode='markers', name='Exit SL ✗',
            marker=dict(color='#ef4444', symbol='x', size=10,
                        line=dict(width=2, color='#7f1d1d')),
        ), secondary_y=True)

    fig.update_layout(
        title=dict(text=title, font=dict(size=13), x=0.5, xanchor='center'),
        height=750,
        hovermode='x unified',
        template='plotly_white',
        dragmode='zoom',
        xaxis=dict(rangeslider=dict(visible=True), fixedrange=False),
        legend=dict(orientation='h', yanchor='bottom', y=1.02, xanchor='center', x=0.5),
    )
    fig.update_yaxes(title_text='Einzel-Equity (USDT)', secondary_y=False, fixedrange=False)
    fig.update_yaxes(title_text='Portfolio-Equity (USDT)', secondary_y=True, fixedrange=False)

    from datetime import datetime as _dt
    ts_stamp = _dt.now().strftime('%Y%m%d_%H%M%S')
    outfile  = f'/tmp/{BOT_NAME}_portfolio_{ts_stamp}.html'
    fig.write_html(outfile)
    print(f'  {G}✓ Chart erstellt: {outfile}{NC}')
    return outfile


def generate_trades_excel(final, capital, start_date, end_date, portfolio_files, strategies_data):
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        print(f'  {Y}openpyxl nicht installiert — Excel übersprungen. (pip install openpyxl){NC}')
        return None

    trades = final.get('trade_history', [])
    if not trades:
        print(f'  {Y}Keine Trades — Excel übersprungen.{NC}')
        return None

    rows = []
    for i, t in enumerate(trades, 1):
        pnl     = t.get('pnl', 0.0)
        outcome = 'TP erreicht' if pnl > 0 else 'SL erreicht'
        sym     = t.get('symbol', '')
        coin    = sym.split('/')[0] if sym else '?'
        rows.append({
            'Nr':              i,
            'Datum (Entry)':   str(t.get('entry_time', ''))[:16].replace('T', ' '),
            'Datum (Exit)':    str(t.get('ts', ''))[:16].replace('T', ' '),
            'Coin':            coin,
            'Timeframe':       t.get('timeframe', ''),
            'Richtung':        t.get('direction', '?').upper(),
            'Entry':           round(t.get('entry', 0.0), 6),
            'Exit':            round(t.get('exit', 0.0), 6),
            'Ergebnis':        outcome,
            'Margin (USDT)':   round(t.get('margin_used', 0.0), 4),
            'PnL (USDT)':      round(pnl, 4),
        })

    # cumulative equity column
    equity = capital
    for r in rows:
        equity += r['PnL (USDT)']
        r['Gesamtkapital'] = round(equity, 4)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Trades'

    header_fill  = PatternFill('solid', fgColor='1E3A5F')
    win_fill     = PatternFill('solid', fgColor='D6F4DC')
    loss_fill    = PatternFill('solid', fgColor='FAD7D7')
    alt_fill     = PatternFill('solid', fgColor='F2F2F2')
    thin_border  = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )
    col_widths = {
        'Nr': 6, 'Datum (Entry)': 18, 'Datum (Exit)': 18, 'Coin': 10,
        'Timeframe': 12, 'Richtung': 10, 'Entry': 14, 'Exit': 14,
        'Ergebnis': 14, 'Margin (USDT)': 16, 'PnL (USDT)': 14, 'Gesamtkapital': 16,
    }

    headers = list(rows[0].keys())
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill      = header_fill
        cell.font      = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col)].width = col_widths.get(h, 14)
    ws.row_dimensions[1].height = 22

    for r_idx, row in enumerate(rows, 2):
        fill = win_fill if row['Ergebnis'] == 'TP erreicht' else (
               loss_fill if r_idx % 2 == 0 else alt_fill)
        for col, key in enumerate(headers, 1):
            cell = ws.cell(row=r_idx, column=col, value=row[key])
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            if key in ('Entry', 'Exit', 'Margin (USDT)', 'PnL (USDT)', 'Gesamtkapital'):
                cell.number_format = '#,##0.0000'
        ws.row_dimensions[r_idx].height = 18

    # Summary block below trades
    pnl_total = final.get('total_pnl_pct', 0)
    wr        = final.get('win_rate', 0)
    dd        = final.get('max_drawdown_pct', 0)
    n         = final.get('trade_count', len(rows))
    eq        = final.get('end_capital', equity)
    sign      = '+' if pnl_total >= 0 else ''

    summary_row = len(rows) + 3
    ws.cell(row=summary_row, column=1, value='Zusammenfassung').font = Font(bold=True, size=11)
    for label, value in [
        ('Trades gesamt',   n),
        ('Win-Rate',        f"{wr:.1f}%"),
        ('PnL',             f"{sign}{pnl_total:.1f}%"),
        ('Final Equity',    f"{eq:.2f} USDT"),
        ('Max Drawdown',    f"{dd:.1f}%"),
        ('Zeitraum',        f"{start_date} → {end_date}"),
    ]:
        ws.cell(row=summary_row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=summary_row, column=2, value=value)
        summary_row += 1

    from datetime import datetime as _dt
    ts_stamp = _dt.now().strftime('%Y%m%d_%H%M%S')
    outfile  = f'/tmp/{BOT_NAME}_trades_{ts_stamp}.xlsx'
    wb.save(outfile)
    print(f'  {G}✓ Excel-Tabelle erstellt: {outfile}{NC}')
    return outfile


def main() -> int:
    parser = argparse.ArgumentParser(description='ZeroBot Portfolio-Optimizer')
    parser.add_argument('--capital',    type=float, default=None)
    parser.add_argument('--max-dd',     type=float, default=30.0)
    parser.add_argument('--start-date', type=str,   default=None)
    parser.add_argument('--end-date',   type=str,   default=None)
    parser.add_argument('--auto-write', action='store_true')
    parser.add_argument('--replot',     action='store_true')
    args = parser.parse_args()

    with open(SETTINGS_PATH) as f:
        settings = json.load(f)
    opt           = settings.get('optimization_settings', {})
    capital       = args.capital or float(opt.get('start_capital', 100))
    max_dd        = args.max_dd
    end_date      = args.end_date   or date.today().strftime('%Y-%m-%d')
    start_date    = args.start_date or (
        date.today() - timedelta(days=DEFAULT_LOOKBACK_DAYS)).strftime('%Y-%m-%d')
    max_positions = int(settings.get('live_trading_settings', {}).get('max_open_positions', 10))

    print(f"\n{'─'*72}")
    print(f"{B}  ZeroBot — Portfolio-Optimizer (EAR){NC}")
    print(f"  Kapital: {capital:.0f} USDT | MaxDD <= {max_dd:.0f}% | "
          f"Zeitraum: {start_date} → {end_date}")
    print(f"{'─'*72}\n")

    config_files = _scan_configs()
    if not config_files:
        print(f"{R}  Keine Configs in {CONFIGS_DIR}{NC}")
        print(f"  → Zuerst den Optimizer ausführen!\n")
        return 1

    print(f"  {len(config_files)} Config(s) gefunden.\n")
    strategies_data = _build_strategies_data(config_files, start_date, end_date)
    if not strategies_data:
        print(f"{R}  Keine Daten geladen.{NC}")
        return 1

    from zerobot.analysis.portfolio_optimizer import run_portfolio_optimizer
    result = run_portfolio_optimizer(capital, strategies_data, start_date, end_date, max_dd)

    if not result or not result.get('optimal_portfolio'):
        print(f"{R}  Kein Portfolio erfüllt MaxDD <= {max_dd:.0f}%.{NC}\n")
        return 0

    portfolio_files = result['optimal_portfolio'][:max_positions]
    final           = result.get('final_result') or {}

    print(f"\n{'='*72}")
    print(f"{B}  Optimales Portfolio — {len(portfolio_files)} Strategie(n){NC}\n")
    for fname in portfolio_files:
        sd = strategies_data.get(fname, {})
        print(f"  {G}✓{NC} {sd.get('symbol', fname):<26} / {sd.get('timeframe', ''):<6}")
    if final:
        pnl = final.get('total_pnl_pct', 0)
        print(f"\n  Endkapital: {final.get('end_capital', 0):.2f} USDT  "
              f"| PnL: {pnl:+.1f}%  | MaxDD: {final.get('max_drawdown_pct', 0):.2f}%")
    print(f"{'='*72}\n")

    labels = [
        f"{strategies_data.get(f, {}).get('symbol', '?')}/{strategies_data.get(f, {}).get('timeframe', '?')}"
        for f in portfolio_files
    ]

    if args.auto_write:
        _write_to_settings(portfolio_files, strategies_data)
        print(f"{G}✓ settings.json aktualisiert — {len(portfolio_files)} Strategie(n).{NC}\n")

        if final:
            pnl = final.get('total_pnl_pct', 0)
            dd  = final.get('max_drawdown_pct', 0)
            n   = final.get('trade_count', 0)
            wr  = final.get('win_rate', 0)
            eq  = final.get('end_capital', 0)
            _send_telegram(
                f"{BOT_NAME} Auto-Optimizer\n"
                f"{len(portfolio_files)} Strategien | {n} Trades | WR: {wr:.1f}%\n"
                f"PnL: {pnl:+.1f}% | MaxDD: {dd:.1f}% | Equity: {eq:.2f} USDT\n"
                f"Zeitraum: {start_date} -> {end_date}")
            html = generate_equity_html(final, capital, start_date, end_date, labels,
                                        portfolio_files, strategies_data)
            if html:
                _send_telegram_doc(html, caption=f'{BOT_NAME} Portfolio-Equity | PnL: {pnl:+.1f}%')
                print(f'  {G}✓ HTML via Telegram gesendet.{NC}')
            excel = generate_trades_excel(final, capital, start_date, end_date,
                                          portfolio_files, strategies_data)
            if excel:
                _send_telegram_doc(excel,
                    caption=f'{BOT_NAME} Trades-Tabelle | {len(portfolio_files)} Strategien | '
                            f'{n} Trades | WR: {wr:.1f}% | Final: {eq:.2f} USDT')
                print(f'  {G}✓ Excel via Telegram gesendet.{NC}')
    else:
        try:
            ans = input("  Portfolio in settings.json eintragen? (j/n): ").strip().lower()
        except (EOFError, KeyboardInterrupt):
            ans = 'n'
        if ans in ('j', 'ja', 'y', 'yes'):
            _write_to_settings(portfolio_files, strategies_data)
            print(f"{G}✓ settings.json aktualisiert.{NC}\n")
        else:
            print(f"{Y}  settings.json NICHT geändert.{NC}\n")

        # Charts + Excel — immer anbieten (unabhängig von settings.json-Entscheidung)
        try:
            chart_ans = input(
                "  Interaktive Charts erstellen & via Telegram senden? (j/n): "
            ).strip().lower()
        except (EOFError, KeyboardInterrupt):
            chart_ans = 'n'

        if chart_ans in ('j', 'ja', 'y', 'yes'):
            if final:
                pnl = final.get('total_pnl_pct', 0)
                n   = final.get('trade_count', 0)
                wr  = final.get('win_rate', 0)
                eq  = final.get('end_capital', 0)
                html = generate_equity_html(final, capital, start_date, end_date, labels,
                                            portfolio_files, strategies_data)
                if html:
                    _send_telegram_doc(html,
                        caption=f'{BOT_NAME} Portfolio-Equity | PnL: {pnl:+.1f}% | '
                                f'Equity: {eq:.2f} USDT | MaxDD: {final.get("max_drawdown_pct", 0):.1f}%')
                    print(f'  {G}✓ HTML via Telegram gesendet.{NC}')
                excel = generate_trades_excel(final, capital, start_date, end_date,
                                              portfolio_files, strategies_data)
                if excel:
                    _send_telegram_doc(excel,
                        caption=f'{BOT_NAME} Trades-Tabelle | {len(portfolio_files)} Strategien | '
                                f'{n} Trades | WR: {wr:.1f}% | Final: {eq:.2f} USDT')
                    print(f'  {G}✓ Excel via Telegram gesendet.{NC}')
            else:
                print(f'  {Y}Keine Simulation-Daten für Charts.{NC}')

    return 0


if __name__ == '__main__':
    sys.exit(main())
